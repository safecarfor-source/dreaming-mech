#!/usr/bin/env python3
"""엑셀 파일 완전 분석 스크립트 - 정비소 현황판"""

import openpyxl
from openpyxl.utils import get_column_letter
from collections import defaultdict
import json
import re

FILE_PATH = "/Users/shinjeayoun/Downloads/(메인)정비소 현황판  (1).xlsx"

def analyze_workbook():
    # data_only=False로 수식 보존
    wb = openpyxl.load_workbook(FILE_PATH, data_only=False)
    # data_only=True로 계산된 값도 읽기
    wb_values = openpyxl.load_workbook(FILE_PATH, data_only=True)

    print("=" * 80)
    print("📊 엑셀 파일 완전 분석 보고서")
    print(f"파일: {FILE_PATH}")
    print("=" * 80)

    # 1. 시트 목록
    print(f"\n{'='*80}")
    print("1. 시트 목록 (총 {0}개)".format(len(wb.sheetnames)))
    print("=" * 80)
    for i, name in enumerate(wb.sheetnames, 1):
        ws = wb[name]
        print(f"  [{i}] '{name}' - 행: {ws.max_row}, 열: {ws.max_column}")

    # 각 시트별 상세 분석
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws_val = wb_values[sheet_name]
        analyze_sheet(ws, ws_val, sheet_name)

def analyze_sheet(ws, ws_val, sheet_name):
    print(f"\n{'#'*80}")
    print(f"# 시트: '{sheet_name}'")
    print(f"# 크기: {ws.max_row}행 x {ws.max_column}열")
    print(f"{'#'*80}")

    # 2. 병합 셀 분석
    merged = list(ws.merged_cells.ranges)
    if merged:
        print(f"\n  --- 병합 셀 ({len(merged)}개) ---")
        for mc in merged[:50]:  # 최대 50개만 표시
            # 병합 범위의 첫 셀 값 가져오기
            first_cell = ws.cell(mc.min_row, mc.min_col)
            val = first_cell.value if first_cell.value else "(빈값)"
            print(f"    {mc} => 값: {val}")
        if len(merged) > 50:
            print(f"    ... 외 {len(merged)-50}개 더 있음")

    # 3. 전체 데이터 스캔 (첫 50행)
    print(f"\n  --- 데이터 미리보기 (최대 50행) ---")
    for row_idx in range(1, min(ws.max_row + 1, 51)):
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.value is not None:
                val_str = str(cell.value)
                if len(val_str) > 30:
                    val_str = val_str[:30] + "..."
                row_data.append(f"{get_column_letter(col_idx)}{row_idx}={val_str}")
        if row_data:
            print(f"    행{row_idx:3d}: {' | '.join(row_data)}")

    # 4. 수식 분석
    formulas = []
    formula_patterns = defaultdict(list)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula = cell.value
                cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                # 계산된 값도 가져오기
                calc_val = ws_val.cell(cell.row, cell.column).value
                formulas.append((cell_ref, formula, calc_val))

                # 수식 패턴 분류
                func_match = re.findall(r'([A-Z]+)\s*\(', formula.upper())
                for func in func_match:
                    formula_patterns[func].append(cell_ref)

    if formulas:
        print(f"\n  --- 수식 ({len(formulas)}개) ---")
        for ref, formula, calc_val in formulas:
            calc_str = f" => 계산값: {calc_val}" if calc_val is not None else ""
            print(f"    {ref}: {formula}{calc_str}")

        print(f"\n  --- 수식 함수별 사용 현황 ---")
        for func, cells in sorted(formula_patterns.items()):
            print(f"    {func}: {len(cells)}회 사용 - 셀: {', '.join(cells[:10])}" +
                  (f" 외 {len(cells)-10}개" if len(cells) > 10 else ""))

    # 5. 셀 서식 분석
    print(f"\n  --- 셀 서식 분석 ---")
    number_formats = defaultdict(list)
    font_colors = set()
    fill_colors = set()
    borders_count = 0

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                # 숫자 포맷
                if cell.number_format and cell.number_format != 'General':
                    nf = cell.number_format
                    ref = f"{get_column_letter(cell.column)}{cell.row}"
                    number_formats[nf].append(ref)

                # 폰트 색상
                if cell.font and cell.font.color and cell.font.color.rgb:
                    font_colors.add(str(cell.font.color.rgb))

                # 배경색
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    rgb = str(cell.fill.fgColor.rgb)
                    if rgb != '00000000':
                        fill_colors.add(rgb)

                # 테두리
                try:
                    if cell.border:
                        sides = [cell.border.left, cell.border.right,
                                 cell.border.top, cell.border.bottom]
                        if any(s and s.style for s in sides):
                            borders_count += 1
                except:
                    pass

    if number_formats:
        print(f"    숫자 포맷:")
        for fmt, cells in number_formats.items():
            print(f"      '{fmt}': {len(cells)}개 셀 (예: {', '.join(cells[:5])})")

    if fill_colors:
        print(f"    배경색: {', '.join(list(fill_colors)[:20])}")
    if font_colors:
        print(f"    글자색: {', '.join(list(font_colors)[:20])}")
    print(f"    테두리 있는 셀: {borders_count}개")

    # 6. 조건부 서식
    if ws.conditional_formatting:
        print(f"\n  --- 조건부 서식 ({len(ws.conditional_formatting)}개) ---")
        for cf in ws.conditional_formatting:
            print(f"    범위: {cf}")
            for rule in cf.rules:
                print(f"      규칙: type={rule.type}, operator={rule.operator}, formula={rule.formula}")
                if rule.dxf:
                    if rule.dxf.fill:
                        print(f"        채우기: {rule.dxf.fill}")
                    if rule.dxf.font:
                        print(f"        글꼴: {rule.dxf.font}")

    # 7. 데이터 유효성 검사
    if ws.data_validations:
        print(f"\n  --- 데이터 유효성 검사 ({len(ws.data_validations.dataValidation)}개) ---")
        for dv in ws.data_validations.dataValidation:
            print(f"    범위: {dv.sqref}")
            print(f"      유형: {dv.type}, 수식1: {dv.formula1}, 수식2: {dv.formula2}")
            if dv.prompt:
                print(f"      안내: {dv.prompt}")
            if dv.error:
                print(f"      오류: {dv.error}")

    # 8. 수동 입력 vs 자동 계산 필드 분석
    print(f"\n  --- 수동 입력 vs 자동 계산 ---")
    manual_cells = 0
    formula_cells = 0
    empty_cells = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is None:
                empty_cells += 1
            elif isinstance(cell.value, str) and cell.value.startswith('='):
                formula_cells += 1
            else:
                manual_cells += 1

    total = manual_cells + formula_cells + empty_cells
    print(f"    수동 입력: {manual_cells}개 셀")
    print(f"    자동 계산(수식): {formula_cells}개 셀")
    print(f"    빈 셀: {empty_cells}개")
    print(f"    전체: {total}개 셀")

    # 9. 행/열 숨김 분석
    hidden_rows = [r for r in range(1, ws.max_row+1)
                   if ws.row_dimensions[r].hidden]
    hidden_cols = [c for c in range(1, ws.max_column+1)
                   if ws.column_dimensions[get_column_letter(c)].hidden]
    if hidden_rows:
        print(f"\n    숨겨진 행: {hidden_rows[:20]}" +
              (f" 외 {len(hidden_rows)-20}개" if len(hidden_rows) > 20 else ""))
    if hidden_cols:
        col_letters = [get_column_letter(c) for c in hidden_cols]
        print(f"    숨겨진 열: {col_letters}")

    # 10. 열 너비
    print(f"\n  --- 열 너비 설정 ---")
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        width = ws.column_dimensions[col_letter].width
        if width:
            print(f"    {col_letter}열: 너비={width}")

if __name__ == "__main__":
    analyze_workbook()
